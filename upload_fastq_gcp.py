# upload_fastq_gcp.py
#!/usr/bin/env python3
# activate gcp-env if required
"""
author @Sharu
This scripts takes command "upload_fastq_gcp" and uploads fastqs to gcp
computes MD5sum for both local and gcp fastqs and gives match status
Optional commands: --validate-only (no Fastq upload), --gcp-folder-md5 (gives MD5Sum of all fastqs in GCP Folder)
Usage: upload_fastq_gcp GCP_folder_path or python3 upload_fastq_gcp.py GCP_Folder_path
"""

import os
import sys
import subprocess
import hashlib
import base64
import binascii
import logging
import argparse

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font


# Get current working directory
CURRENT_DIR = os.getcwd()

# LOGGING
# Creates both terminal and log-file based logging
LOG_FILE = os.path.join(
    CURRENT_DIR,
    "upload_fastq_gcp.log"
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(sys.stdout)
    ]
)

# FUNCTIONS

# Detect all .fastq.gz files from current directory
def get_fastq_files():

    fastqs = sorted([
        f for f in os.listdir(".")
        if f.endswith(".fastq.gz")
    ])

    return fastqs

# Create a text file containing absolute FASTQ paths
def create_fastq_list(fastqs):

    with open("fastq_list.txt", "w") as f:

        for fq in fastqs:
            f.write(os.path.abspath(fq) + "\n")

    logging.info(
        f"Created fastq_list.txt with {len(fastqs)} FASTQs"
    )

# Calculate local MD5 checksum for a FASTQ file
def calculate_md5(file_path):

    md5_hash = hashlib.md5()

    with open(file_path, "rb") as f:

        for chunk in iter(lambda: f.read(8192), b""):
            md5_hash.update(chunk)

    return md5_hash.hexdigest()

# Fetch MD5 checksum of a FASTQ file stored in GCP bucket
def get_gcp_md5(gcp_path):

    try:

        result = subprocess.check_output(
            f'gsutil stat "{gcp_path}"',
            shell=True,
            text=True
        )

        for line in result.splitlines():

            if "Hash (md5)" in line:

                md5_base64 = line.split(":")[1].strip()

                md5_bytes = base64.b64decode(md5_base64)

                md5_hex = binascii.hexlify(
                    md5_bytes
                ).decode()

                return md5_hex

    except subprocess.CalledProcessError:

        return None

    return None

# Upload all local FASTQ files to specified GCP folder
def upload_fastqs(fastqs, gcp_folder):

    logging.info("Starting upload")

    for fq in fastqs:

        logging.info(f"Uploading: {fq}")

        cmd = (
            f'gsutil -m cp "{fq}" "{gcp_folder}/"'
        )

        ret = os.system(cmd)

        if ret != 0:

            logging.error(
                f"Upload failed: {fq}"
            )

            sys.exit(1)

        logging.info(
            f"Upload completed: {fq}"
        )

# Group FASTQs into R1/R2 pairs using sample name
def group_fastqs(fastqs):

    sample_dict = {}

    for fq in fastqs:

        if "_R1" in fq:

            sample = fq.split("_R1")[0]

            sample_dict.setdefault(
                sample,
                {}
            )["R1"] = fq

        elif "_R2" in fq:

            sample = fq.split("_R2")[0]

            sample_dict.setdefault(
                sample,
                {}
            )["R2"] = fq

    return sample_dict

# Create Excel report containing local vs GCP MD5 validation results
def create_validation_excel(results):

    wb = Workbook()

    ws = wb.active

    ws.title = "FASTQ_MD5_Validation"

    headers = [
        "fastq_name",
        "local_MD5sum",
        "GCP_MD5sum",
        "match_status"
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in results:
        ws.append(row)

    output_file = (
        f"FASTQ_GCP_MD5_Report_"
        f"{datetime.now().strftime('%d_%b_%Y')}.xlsx"
    )

    wb.save(output_file)

    logging.info(
        f"Created Excel: {output_file}"
    )

    return output_file

# Compare local and GCP MD5 values for all FASTQs in local folder
def validate_fastqs(sample_dict, gcp_folder):

    logging.info("Starting validation")

    results = []

    for sample, reads in sorted(
        sample_dict.items()
    ):

        logging.info(
            f"Processing sample: {sample}"
        )

        # ---------------- R1 ----------------

        r1 = reads.get("R1")

        if r1:

            local_md5_r1 = calculate_md5(r1)

            gcp_r1_path = (
                f"{gcp_folder}/"
                f"{os.path.basename(r1)}"
            )

            gcp_md5_r1 = get_gcp_md5(
                gcp_r1_path
            )

            status_r1 = (
                "MATCH"
                if local_md5_r1 == gcp_md5_r1
                else "FAIL"
            )

            results.append([
                os.path.basename(r1),
                local_md5_r1,
                gcp_md5_r1,
                status_r1
            ])

            logging.info(
                f"{os.path.basename(r1)} | "
                f"LOCAL={local_md5_r1} | "
                f"GCP={gcp_md5_r1} | "
                f"STATUS={status_r1}"
            )

        # ---------------- R2 ----------------

        r2 = reads.get("R2")

        if r2:

            local_md5_r2 = calculate_md5(r2)

            gcp_r2_path = (
                f"{gcp_folder}/"
                f"{os.path.basename(r2)}"
            )

            gcp_md5_r2 = get_gcp_md5(
                gcp_r2_path
            )

            status_r2 = (
                "MATCH"
                if local_md5_r2 == gcp_md5_r2
                else "FAIL"
            )

            results.append([
                os.path.basename(r2),
                local_md5_r2,
                gcp_md5_r2,
                status_r2
            ])

            logging.info(
                f"{os.path.basename(r2)} | "
                f"LOCAL={local_md5_r2} | "
                f"GCP={gcp_md5_r2} | "
                f"STATUS={status_r2}"
            )

    return results

# List all FASTQ files present inside GCP folder

def list_gcp_fastqs(gcp_folder):

    cmd = (
        f'gsutil ls '
        f'"{gcp_folder}/*.fastq.gz"'
    )

    result = subprocess.check_output(
        cmd,
        shell=True,
        text=True
    )

    fastqs = result.strip().splitlines()

    return fastqs

# Generate Excel report containing MD5 values of GCP FASTQs

def gcp_folder_md5_report(gcp_folder):

    logging.info(
        "Running GCP folder MD5 mode"
    )

    gcp_fastqs = list_gcp_fastqs(
        gcp_folder
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "GCP_FOLDER_MD5"

    headers = [
        "fastq_name",
        "GCP_MD5sum"
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    print(
        f"\nTotal FASTQs in GCP: "
        f"{len(gcp_fastqs)}\n"
    )

    for gcp_path in sorted(gcp_fastqs):

        fq_name = os.path.basename(
            gcp_path
        )

        print(
            f"Fetching MD5: {fq_name}"
        )

        gcp_md5 = get_gcp_md5(
            gcp_path
        )

        ws.append([
            fq_name,
            gcp_md5
        ])

        logging.info(
            f"{fq_name} | "
            f"MD5={gcp_md5}"
        )

    output_file = (
        f"GCP_FOLDER_MD5_Report_"
        f"{datetime.now().strftime('%d_%b_%Y')}.xlsx"
    )

    wb.save(output_file)

    print(
        f"\nReport Created: "
        f"{output_file}\n"
    )

    logging.info(
        f"Created: {output_file}"
    )

# MAIN
def main():

    parser = argparse.ArgumentParser(
        description=(
            "Upload FASTQs to GCP "
            "and validate MD5"
        )
    )

    parser.add_argument(
        "gcp_folder",
        help="Destination GCP folder"
    )

    parser.add_argument(
        "--validate-only",
        action="store_true",
        help=(
            "Skip upload and "
            "only validate local vs GCP MD5"
        )
    )

    parser.add_argument(
        "--gcp-folder-md5",
        action="store_true",
        help=(
            "Fetch MD5 of all FASTQs "
            "present in GCP folder"
        )
    )

    args = parser.parse_args()

    gcp_folder = args.gcp_folder.rstrip("/")

    validate_only = args.validate_only

    # GCP FOLDER MD5 MODE

    if args.gcp_folder_md5:

        gcp_folder_md5_report(
            gcp_folder
        )

        print(
            "\nGCP FOLDER MD5 "
            "COMPLETED\n"
        )

        sys.exit(0)

    # GET LOCAL FASTQS

    fastqs = get_fastq_files()

    if len(fastqs) == 0:

        logging.error(
            "No FASTQ files found"
        )

        sys.exit(1)

    print("\nDetected FASTQs:\n")

    for fq in fastqs:
        print(fq)

    print(
        f"\nTotal FASTQs: "
        f"{len(fastqs)}"
    )

    logging.info(
        f"Total FASTQs: {len(fastqs)}"
    )

    create_fastq_list(fastqs)

    # VALIDATE ONLY MODE (No fastq upload)

    if validate_only:

        print(
            "\nVALIDATE ONLY MODE ENABLED"
        )

        print(
            "Skipping upload\n"
        )

        logging.info(
            "Running VALIDATE ONLY mode"
        )

    else:

        confirm = input(
            "\nProceed with upload? "
            "Type YES to continue: "
        )

        if confirm.strip().upper() != "YES":

            logging.warning(
                "Upload cancelled by user"
            )

            print(
                "\nUpload cancelled\n"
            )

            sys.exit(0)

        upload_fastqs(
            fastqs,
            gcp_folder
        )

    # VALIDATION

    sample_dict = group_fastqs(
        fastqs
    )

    results = validate_fastqs(
        sample_dict,
        gcp_folder
    )

    excel_file = create_validation_excel(
        results
    )

    # ========================================================

    print("\n================================")
    print("PROCESS COMPLETED")
    print("================================")
    print(f"Excel Report : {excel_file}")
    print(f"Log File     : {LOG_FILE}")
    print("================================\n")

    logging.info(
        "Process completed successfully"
    )

# ============================================================

if __name__ == "__main__":
    main()
